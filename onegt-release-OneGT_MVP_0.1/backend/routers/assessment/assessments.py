"""
Assessment Router - Ported from Assessment-Portal-1/backend/main.py.
Handles assessment creation, retrieval, and grading.
"""
import json
import random
import logging
from datetime import datetime
from typing import List, Optional, Any
from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File, Form
from fastapi.responses import JSONResponse

from middleware.auth_middleware import get_current_user, TokenData
from utils.assessment_db import (
    save_assessment, get_assessment, update_assessment, delete_assessment,
    get_all_assessments, save_result, get_user_attempt_count, mark_assessment_started,
    upsert_user_from_token, _gen_id, _now_iso
)
from services.assessment_ai import generate_questions

logger = logging.getLogger("chrms.assessment.assessments")

router = APIRouter()

@router.post("/create")
async def create_assessment_route(
    request: Request,
    title: str = Form("Untitled Assessment"),
    description: str = Form(""),
    createdBy: str = Form(""),
    assignedTo: str = Form("[]"),
    scheduledFrom: str = Form(""),
    scheduledTo: str = Form(""),
    durationMinutes: str = Form("30"),
    timePerQuestion: str = Form("0"),
    difficulty: str = Form("medium"),
    prompt: str = Form(""),
    questions_json: str = Form("[]", alias="questions"),
    file: UploadFile | None = File(None),
    current_user: TokenData = Depends(get_current_user)
):
    # Sync user
    upsert_user_from_token(current_user)
    
    questions = []
    if questions_json and questions_json != "[]":
        try:
            questions = json.loads(questions_json)
        except Exception as e:
            logger.error(f"Failed to parse questions_json: {e}")

    if not questions:
        file_content = ""
        if file and file.filename:
            raw = await file.read()
            fname = file.filename.lower()

            if fname.endswith(".pdf"):
                try:
                    from PyPDF2 import PdfReader
                    import io as _io
                    reader = PdfReader(_io.BytesIO(raw))
                    file_content = "\n".join(p.extract_text() or "" for p in reader.pages)
                    if not file_content.strip():
                        raise HTTPException(400, "The PDF appears to be empty or scanned.")
                except HTTPException:
                    raise
                except Exception as e:
                    raise HTTPException(400, f"PDF parsing failed: {e}")

            elif fname.endswith((".xlsx", ".xls")):
                try:
                    from openpyxl import load_workbook
                    import io as _io
                    wb = load_workbook(_io.BytesIO(raw), read_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if rows:
                        header = [str(c or "") for c in rows[0]]
                        lines = [",".join(header)]
                        for row in list(rows[1:]):
                            lines.append(",".join(str(c or "") for c in row))
                        file_content = "\n".join(lines)
                except Exception as e:
                    raise HTTPException(400, f"Excel read failed: {e}")
            else:
                try:
                    file_content = raw.decode("utf-8", errors="replace")
                except:
                    file_content = str(raw)

        questions = generate_questions(prompt, file_content or None)

    if not questions:
        raise HTTPException(400, "No questions could be extracted or generated.")

    random.shuffle(questions)
    if len(questions) > 150:
        questions = questions[:150]

    try:
        tpq = int(timePerQuestion)
        if tpq > 0:
            for q in questions:
                q["time_limit_seconds"] = tpq
    except:
        pass

    assessment_id = _gen_id()
    try:
        assigned = json.loads(assignedTo)
    except:
        assigned = []

    save_assessment({
        "assessment_id": assessment_id,
        "title": title,
        "description": description,
        "difficulty": difficulty,
        "questions": questions,
        "created_by": current_user.associate_id, # Override with actual user
        "created_at": _now_iso(),
        "scheduled_from": scheduledFrom or None,
        "scheduled_to": scheduledTo or None,
        "duration_minutes": int(durationMinutes),
        "assigned_to": assigned,
    })

    return {"assessment_id": assessment_id, "question_count": len(questions)}

@router.get("/{assessment_id}")
def get_assessment_route(
    assessment_id: str, 
    current_user: TokenData = Depends(get_current_user)
):
    upsert_user_from_token(current_user)
    
    assessment = get_assessment(assessment_id)
    if not assessment:
        raise HTTPException(404, "Assessment not found")

    user_id = current_user.associate_id
    assigned_to = assessment.get("assigned_to", [])
    
    # Check if user has access (Admin, Creator, or Assigned Candidate)
    is_authorized = (current_user.role == "Admin" or 
                     assessment.get("created_by") == user_id or 
                     user_id in assigned_to)
    
    if not is_authorized:
        raise HTTPException(403, "You are not authorized to access this assessment.")

    if user_id in assigned_to and current_user.role == "Associate":
        attempts = get_user_attempt_count(assessment_id, user_id)
        retake = assessment.get("retake_permissions", [])
        if attempts > 0 and user_id not in retake:
            return JSONResponse(
                {"error": "You have already attempted this assessment.",
                 "already_attempted": True, "attempt_count": attempts},
                status_code=403,
            )

    # Strip correct answers if it's a candidate
    safe_qs = []
    is_candidate = (current_user.role == "Associate")
    
    for q in assessment.get("questions", []):
        if is_candidate:
            sq = {k: v for k, v in q.items() if k not in ("correct_option_id", "explanation")}
        else:
            sq = q
        safe_qs.append(sq)

    return {
        "assessment_id": assessment.get("assessment_id"),
        "title": assessment["title"],
        "description": assessment.get("description"),
        "questions": safe_qs,
        "duration_minutes": assessment.get("duration_minutes"),
        "instructions": "Select one option per question. Submit answers as an array of {question_id, option_id}.",
    }

@router.post("/{assessment_id}/start")
async def start_assessment_route(
    assessment_id: str, 
    current_user: TokenData = Depends(get_current_user)
):
    upsert_user_from_token(current_user)
    
    assessment = get_assessment(assessment_id)
    if not assessment:
        raise HTTPException(404, "Assessment not found")

    mark_assessment_started(assessment_id, current_user.associate_id)
    return {"success": True, "message": "Assessment started"}

@router.post("/{assessment_id}/grade")
async def grade_assessment_route(
    assessment_id: str, 
    request: Request,
    current_user: TokenData = Depends(get_current_user)
):
    upsert_user_from_token(current_user)
    
    body = await request.json()
    assessment = get_assessment(assessment_id)
    if not assessment:
        raise HTTPException(404, "Assessment not found")

    user_id = current_user.associate_id
    attempts = get_user_attempt_count(assessment_id, user_id)
    retake_perms = assessment.get("retake_permissions", [])
    has_retake = user_id in retake_perms

    if attempts > 0 and not has_retake:
        raise HTTPException(403, "You have already attempted this assessment.")

    if attempts > 0 and has_retake:
        updated = [uid for uid in retake_perms if uid != user_id]
        update_assessment(assessment_id, {"retake_permissions": updated})

    correct_count: int = 0
    detailed = []

    for q in assessment["questions"]:
        answers = body.get("submissions") or body.get("answers") or []
        user_answer = next((a for a in answers if a["question_id"] == q["id"]), None)
        submitted = user_answer["option_id"] if user_answer else ""
        is_correct = submitted == q.get("correct_option_id", "")
        if is_correct:
            correct_count += 1
        detailed.append({
            "question_id": q["id"],
            "submitted": submitted,
            "correct": q.get("correct_option_id", ""),
            "is_correct": is_correct,
            "points_awarded": 1 if is_correct else 0,
            "explanation": q.get("explanation"),
        })

    total = len(assessment["questions"])
    score = (float(correct_count) / total) * 100 if total else 0.0

    try:
        t_start = datetime.fromisoformat(body["time_started"].replace("Z", "+00:00")).timestamp()
        t_end = datetime.fromisoformat(body["time_submitted"].replace("Z", "+00:00")).timestamp()
    except:
        t_start = 0
        t_end = 0
        
    time_taken = max(0.0, float(t_end) - float(t_start))

    result = {
        "assessment_id": assessment_id,
        "user_id": user_id,
        "score": score,
        "max_score": 100,
        "total_questions": total,
        "correct_count": correct_count,
        "detailed": detailed,
        "analytics": {
            "time_taken_seconds": time_taken,
            "accuracy_percent": score,
            "avg_time_per_question_seconds": time_taken / total if total else 0,
        },
        "graded_at": _now_iso(),
        "tab_switch_count": body.get("tab_switch_count", 0),
        "termination_reason": body.get("termination_reason"),
    }

    save_result(result)
    return result

@router.delete("/{assessment_id}/delete")
def delete_assessment_route(
    assessment_id: str,
    current_user: TokenData = Depends(get_current_user)
):
    if current_user.role not in ["Admin", "Project Manager"]:
        raise HTTPException(403, "Only admins or managers can delete assessments.")
        
    assessment = get_assessment(assessment_id)
    if not assessment:
        raise HTTPException(404, "Assessment not found")
        
    delete_assessment(assessment_id)
    return {"success": True, "message": "Assessment deleted successfully"}
