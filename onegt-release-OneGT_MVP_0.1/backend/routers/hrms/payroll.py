from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from typing import List, Optional
import io
import logging
from services.google_sheets import sheets_service
from models.hrms.payroll import (
    Payroll, PayrollCreate, payroll_to_row, row_to_payroll,
    MONTH_ABBREV_TO_FULL, normalize_month
)
from config import settings

logger = logging.getLogger("chrms.payroll")

router = APIRouter()

@router.get("/", response_model=List[Payroll])
async def get_payroll(
    year: Optional[int] = None,
    month: Optional[str] = None,
    associate_id: Optional[str] = None,
    department: Optional[str] = None
):
    """Get payroll records with optional filters."""
    try:
        records = sheets_service.get_all_records(settings.PAYROLL_SHEET)
        payrolls = []
        
        for idx, r in enumerate(records):
            # Check for employee code in either column name
            emp_code = r.get("Employee Code") or r.get("Associate ID")
            if not emp_code:
                continue
            
            # Get year from either column name
            record_year = r.get("Year") or r.get("Payroll Year", 0)
            try:
                record_year = int(record_year or 0)
            except:
                record_year = 0
            
            # Get month from either column name and normalize
            record_month = str(r.get("Month", "") or r.get("Payroll Month", ""))
            record_month_full = normalize_month(record_month)
            
            # Apply filters
            if year and record_year != year:
                continue
            if month and record_month_full.lower() != month.lower():
                continue
            if associate_id and emp_code != associate_id:
                continue
            if department and r.get("Department Name", "").lower() != department.lower():
                continue
            
            payrolls.append(row_to_payroll(r, idx + 2))
        
        return payrolls
    except Exception as e:
        logger.error(f"Error fetching payroll: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/", response_model=dict)
async def create_payroll(payroll: PayrollCreate):
    """Add a payroll record."""
    try:
        row = payroll_to_row(payroll)
        result = sheets_service.append_row(settings.PAYROLL_SHEET, row)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/bulk", response_model=dict)
async def bulk_create_payroll(payrolls: List[PayrollCreate]):
    """Add multiple payroll records."""
    try:
        count = 0
        for payroll in payrolls:
            row = payroll_to_row(payroll)
            sheets_service.append_row(settings.PAYROLL_SHEET, row)
            count += 1
        
        return {"success": True, "message": f"Added {count} payroll records"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/upload", response_model=dict)
async def upload_payroll_excel(file: UploadFile = File(...)):
    """
    Upload payroll data from Excel file.
    Expects headers starting from row 4: S.No, Employee Code, Employee Name, 
    Date of Joining, Department Name, Designation Name, Category Name,
    Earnings, Statutories Amount, Income Tax, Deductions, Net Pay
    
    Year and Month are extracted from the report metadata in row 3.
    """
    try:
        import openpyxl
        from datetime import datetime
        
        # Read the file
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        
        # Extract period info from row 3 (Report Generated Date: ... | Period: January-2026)
        period_cell = None
        for cell in ws[3]:
            if cell.value and "Period" in str(cell.value):
                period_cell = str(cell.value)
                break
        
        # Parse period (e.g., "Report Generated Date: 02/02/2026 20:58 | Period :January-2026")
        year = datetime.now().year
        month = ""
        
        if period_cell:
            # Try to extract from "Period :Month-Year" or "Period: Month-Year"
            if "Period" in period_cell:
                period_part = period_cell.split("Period")[-1].strip()
                period_part = period_part.replace(":", "").strip()
                if "-" in period_part:
                    parts = period_part.split("-")
                    month = parts[0].strip()
                    try:
                        year = int(parts[1].strip())
                    except:
                        pass
        
        # Find header row (row 4 based on screenshot)
        header_row = 4
        headers = [str(cell.value or "").strip() for cell in ws[header_row]]
        
        # Map expected columns to indices
        col_map = {}
        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if "employee code" in header_lower:
                col_map["employee_code"] = idx
            elif "employee name" in header_lower:
                col_map["employee_name"] = idx
            elif "date of joining" in header_lower:
                col_map["date_of_joining"] = idx
            elif "department" in header_lower:
                col_map["department"] = idx
            elif "designation" in header_lower:
                col_map["designation"] = idx
            elif "category" in header_lower:
                col_map["category"] = idx
            elif "earnings" in header_lower:
                col_map["earnings"] = idx
            elif "statutories" in header_lower:
                col_map["statutories"] = idx
            elif "income tax" in header_lower:
                col_map["income_tax"] = idx
            elif "deductions" in header_lower:
                col_map["deductions"] = idx
            elif "net pay" in header_lower:
                col_map["net_pay"] = idx
        
        # Read data rows
        records_added = 0
        records_skipped = 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
            # Get employee code
            emp_code_idx = col_map.get("employee_code")
            if emp_code_idx is None:
                continue
            
            emp_code = row[emp_code_idx].value
            if not emp_code:
                continue
            
            # Extract all values
            def get_val(key, default=""):
                idx = col_map.get(key)
                if idx is not None and idx < len(row):
                    val = row[idx].value
                    return val if val is not None else default
                return default
            
            def get_float(key, default=0):
                val = get_val(key, default)
                try:
                    return float(val or 0)
                except:
                    return default
            
            # Format date of joining
            doj = get_val("date_of_joining", "")
            if hasattr(doj, "strftime"):
                doj = doj.strftime("%Y-%m-%d")
            else:
                doj = str(doj) if doj else ""
            
            # Create payroll record
            payroll_data = PayrollCreate(
                payroll_month=month,
                payroll_year=year,
                associate_id=str(emp_code),
                associate_name=str(get_val("employee_name", "")),
                date_of_joining=doj,
                department_name=str(get_val("department", "")),
                designation_name=str(get_val("designation", "")),
                category_name=str(get_val("category", "")),
                earnings=get_float("earnings"),
                statutories_amount=get_float("statutories"),
                income_tax=get_float("income_tax"),
                deductions=get_float("deductions"),
                net_pay=get_float("net_pay")
            )
            
            # Append to sheet
            row_data = payroll_to_row(payroll_data)
            sheets_service.append_row(settings.PAYROLL_SHEET, row_data)
            records_added += 1
        
        return {
            "success": True,
            "message": f"Successfully uploaded {records_added} payroll records for {month} {year}",
            "records_added": records_added,
            "records_skipped": records_skipped,
            "period": f"{month} {year}"
        }
        
    except ImportError:
        raise HTTPException(
            status_code=500, 
            detail="openpyxl library not installed. Run: pip install openpyxl"
        )
    except Exception as e:
        logger.error(f"Error uploading payroll: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/{row_index}", response_model=dict)
async def delete_payroll(row_index: int):
    """Delete a payroll record by row index."""
    try:
        result = sheets_service.delete_row(settings.PAYROLL_SHEET, row_index)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/summary", response_model=dict)
async def get_payroll_summary(
    year: int = Query(..., description="Year"),
    month: Optional[str] = None
):
    """Get payroll summary statistics."""
    try:
        records = sheets_service.get_all_records(settings.PAYROLL_SHEET)
        
        total_earnings = 0
        total_deductions = 0
        total_net_pay = 0
        employee_count = 0
        department_totals = {}
        
        for r in records:
            emp_code = r.get("Employee Code") or r.get("Associate ID")
            if not emp_code:
                continue
            
            record_year = r.get("Year") or r.get("Payroll Year", 0)
            try:
                record_year = int(record_year or 0)
            except:
                record_year = 0
            
            if record_year != year:
                continue
            
            if month:
                record_month = str(r.get("Month", "") or r.get("Payroll Month", ""))
                record_month_full = normalize_month(record_month)
                if record_month_full.lower() != month.lower():
                    continue
            
            employee_count += 1
            earnings = float(r.get("Earnings", 0) or 0)
            deductions = float(r.get("Deductions", 0) or 0) + float(r.get("Income Tax", 0) or 0)
            net_pay = float(r.get("Net Pay", 0) or 0)
            
            total_earnings += earnings
            total_deductions += deductions
            total_net_pay += net_pay
            
            dept = r.get("Department Name", "Unknown")
            if dept not in department_totals:
                department_totals[dept] = {"count": 0, "total_pay": 0}
            department_totals[dept]["count"] += 1
            department_totals[dept]["total_pay"] += net_pay
        
        return {
            "year": year,
            "month": month,
            "employee_count": employee_count,
            "total_earnings": total_earnings,
            "total_deductions": total_deductions,
            "total_net_pay": total_net_pay,
            "department_breakdown": department_totals
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/associate/{associate_id}/history", response_model=List[Payroll])
async def get_associate_payroll_history(associate_id: str):
    """Get payroll history for an associate."""
    try:
        records = sheets_service.get_all_records(settings.PAYROLL_SHEET)
        payrolls = []
        
        for idx, r in enumerate(records):
            emp_code = r.get("Employee Code") or r.get("Associate ID")
            if emp_code == associate_id:
                payrolls.append(row_to_payroll(r, idx + 2))
        
        # Sort by year and month
        month_order = {
            "january": 1, "february": 2, "march": 3, "april": 4,
            "may": 5, "june": 6, "july": 7, "august": 8,
            "september": 9, "october": 10, "november": 11, "december": 12
        }
        
        payrolls.sort(
            key=lambda p: (p.payroll_year, month_order.get(p.payroll_month.lower(), 0)),
            reverse=True
        )
        
        return payrolls
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
